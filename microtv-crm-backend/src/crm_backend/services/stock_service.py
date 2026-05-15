"""Casos de uso del flujo inicial real de depósito."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO, StringIO
import re
from pathlib import PurePath

from sqlalchemy import select

from crm_backend.core.config import Settings
from crm_backend.core.exceptions import (
    DuplicateStockProductCodeError,
    InvalidStockQuantityError,
    InventoryAdminRequiredError,
    InventoryAccessDeniedError,
    StockBackupNotFoundError,
    StockCategoryNotFoundError,
    StockImportConflictError,
    StockImportNotFoundError,
    StockImportValidationError,
    StockProductNotFoundError,
)
from crm_backend.models import StockBackup, StockBackupRow, StockCategory, StockImportBatch, StockImportRow, StockLevel, StockProduct
from crm_backend.models.notification import NotificationEntityType, NotificationType
from crm_backend.repositories import CrmUserRepository, StockCategoryRepository, StockProductRepository
from crm_backend.schemas import (
    StockBackupStatusResponse,
    StockImportConfirmResponse,
    StockImportPreviewResponse,
    StockImportRowPreview,
    StockProductResponse,
    StockRollbackResponse,
)
from crm_backend.services.activity_log_service import ActivityLogService
from crm_backend.services.auth_service import ResolvedCrmSession
from crm_backend.services.notification_service import NotificationService
from crm_backend.services.permission_service import (
    PERMISSION_STOCK_DELETE_PRODUCT,
    PERMISSION_STOCK_MANAGE,
    PermissionService,
)


@dataclass(slots=True)
class CreateStockProductCommand:
    """Agrupa los datos de alta de producto.

    Attributes:
        name: Nombre del producto.
        product_code: Código visible obligatorio del producto.
        category_id: Categoría elegida.
        initial_stock: Stock inicial.
        image_url: URL opcional de imagen.
    """

    name: str
    product_code: str
    category_id: str
    initial_stock: int
    image_url: str | None
    requires_tracking: bool
    minimum_stock: int = 3


class StockApplicationService:
    """Orquesta el módulo inicial real de depósito."""

    def __init__(
        self,
        settings: Settings,
        category_repository: StockCategoryRepository,
        product_repository: StockProductRepository,
        notification_service: NotificationService | None = None,
        user_repository: CrmUserRepository | None = None,
        permission_service: PermissionService | None = None,
        activity_log_service: ActivityLogService | None = None,
    ) -> None:
        """Crea el servicio.

        Args:
            settings: Configuración de la aplicación.
            category_repository: Repositorio de categorías.
            product_repository: Repositorio de productos.
        """

        self._settings = settings
        self._category_repository = category_repository
        self._product_repository = product_repository
        self._notification_service = notification_service
        self._user_repository = user_repository
        self._permission_service = permission_service
        self._activity_log_service = activity_log_service

    def list_categories(self, actor: ResolvedCrmSession) -> list[StockCategory]:
        """Lista categorías disponibles para el actor autenticado.

        Args:
            actor: Sesión CRM autenticada.

        Returns:
            list[StockCategory]: Categorías activas.
        """

        self._ensure_inventory_read_access(actor)
        return self._category_repository.list_active()

    def list_products(self, actor: ResolvedCrmSession) -> list[StockProduct]:
        """Lista productos activos para el actor autenticado.

        Args:
            actor: Sesión CRM autenticada.

        Returns:
            list[StockProduct]: Productos activos.
        """

        self._ensure_inventory_read_access(actor)
        return self._product_repository.list_active()

    def create_product(self, actor: ResolvedCrmSession, command: CreateStockProductCommand) -> StockProduct:
        """Crea un producto real en depósito.

        Args:
            actor: Sesión CRM autenticada.
            command: Datos del producto.

        Returns:
            StockProduct: Producto persistido.
        """

        self._ensure_inventory_write_access(actor)
        category = self._category_repository.get_active_by_id(command.category_id)
        if category is None:
            raise StockCategoryNotFoundError()
        if self._product_repository.get_by_code(command.product_code) is not None:
            raise DuplicateStockProductCodeError()

        warehouse_id = self._product_repository.get_default_warehouse_id()
        product = StockProduct.create(
            name=command.name,
            product_code=command.product_code,
            stock_category_id=category.stock_category_id,
            initial_stock=command.initial_stock,
            image_url=command.image_url,
            requires_tracking=command.requires_tracking,
            minimum_stock=command.minimum_stock,
            actor_crm_user_id=actor.crm_user.crm_user_id,
            warehouse_id=warehouse_id,
        )
        saved = self._product_repository.save(product)
        self._log_event(
            "stock.product_created",
            actor,
            saved.product_id,
            saved.product_name,
            {"product_code": saved.visible_product_code, "stock": saved.current_stock},
        )
        return saved

    def increase_stock(self, actor: ResolvedCrmSession, product_id: str, quantity: int) -> StockProduct:
        """Aumenta el stock de un producto existente.

        Args:
            actor: Sesión CRM autenticada.
            product_id: Producto afectado.
            quantity: Cantidad a sumar.

        Returns:
            StockProduct: Producto actualizado.
        """

        self._ensure_inventory_write_access(actor)
        product = self._get_operable_product(product_id)
        product.increase_stock(
            quantity=quantity,
            actor_crm_user_id=actor.crm_user.crm_user_id,
            warehouse_id=self._product_repository.get_default_warehouse_id(),
        )
        saved = self._product_repository.save(product)
        self._log_event(
            "stock.movement",
            actor,
            saved.product_id,
            saved.product_name,
            {"kind": "increase", "quantity": quantity, "stock": saved.current_stock},
        )
        return saved

    def decrease_stock(self, actor: ResolvedCrmSession, product_id: str, quantity: int) -> StockProduct:
        """Disminuye el stock de un producto existente.

        Args:
            actor: Sesión CRM autenticada.
            product_id: Producto afectado.
            quantity: Cantidad a descontar.

        Returns:
            StockProduct: Producto actualizado.
        """

        self._ensure_inventory_write_access(actor)
        product = self._get_operable_product(product_id)
        product.decrease_stock(
            quantity=quantity,
            actor_crm_user_id=actor.crm_user.crm_user_id,
            warehouse_id=self._product_repository.get_default_warehouse_id(),
        )
        saved_product = self._product_repository.save(product)
        self._log_event(
            "stock.movement",
            actor,
            saved_product.product_id,
            saved_product.product_name,
            {"kind": "decrease", "quantity": quantity, "stock": saved_product.current_stock},
        )

        if self._notification_service is not None and self._user_repository is not None:
            stock_now = saved_product.current_stock
            if stock_now == 0:
                notification_type = NotificationType.STOCK_OUT
                title = f"Sin stock: {saved_product.visible_product_code}"
                body = f"El producto '{saved_product.product_name}' llegó a 0 unidades."
            elif stock_now < saved_product.minimum_stock:
                notification_type = NotificationType.STOCK_LOW
                title = f"Stock bajo: {saved_product.visible_product_code} ({stock_now} unidades)"
                body = f"El producto '{saved_product.product_name}' tiene menos de {saved_product.minimum_stock} unidades disponibles."
            else:
                notification_type = None

            if notification_type is not None:
                deposito_ids = [user.crm_user_id for user in self._user_repository.list_active_by_role_key("deposito")]
                ejecutivo_ids = [user.crm_user_id for user in self._user_repository.list_active_by_role_key("ejecutivo")]
                recipient_ids = list({*deposito_ids, *ejecutivo_ids})
                self._notification_service.notify_bulk(
                    recipient_crm_user_ids=recipient_ids,
                    notification_type=notification_type,
                    title=title,
                    body=body,
                    entity_type=NotificationEntityType.STOCK_PRODUCT,
                    entity_id=saved_product.product_id,
                )

        return saved_product

    def update_product_location(
        self,
        actor: ResolvedCrmSession,
        product_id: str,
        shelf_id: str,
        shelf_height: int,
    ) -> StockProduct:
        self._ensure_inventory_write_access(actor)
        product = self._get_operable_product(product_id)
        product.shelf_id = shelf_id
        product.shelf_height = shelf_height
        saved = self._product_repository.save(product)
        self._log_event(
            "stock.product_updated",
            actor,
            saved.product_id,
            saved.product_name,
            {"shelf_id": shelf_id, "shelf_height": shelf_height},
        )
        return saved

    def set_stock(
        self,
        actor: ResolvedCrmSession,
        product_id: str,
        quantity: int,
    ) -> StockProduct:
        self._ensure_inventory_write_access(actor)
        product = self._get_operable_product(product_id)

        current = product.current_stock
        if quantity > current:
            product.increase_stock(
                quantity=quantity - current,
                actor_crm_user_id=actor.crm_user.crm_user_id,
                warehouse_id=self._product_repository.get_default_warehouse_id(),
            )
        elif quantity < current:
            product.decrease_stock(
                quantity=current - quantity,
                actor_crm_user_id=actor.crm_user.crm_user_id,
                warehouse_id=self._product_repository.get_default_warehouse_id(),
            )

        saved_product = self._product_repository.save(product)
        self._log_event(
            "stock.movement",
            actor,
            saved_product.product_id,
            saved_product.product_name,
            {"kind": "set", "quantity": quantity, "stock": saved_product.current_stock},
        )

        if self._notification_service is not None and self._user_repository is not None:
            stock_now = saved_product.current_stock
            if stock_now == 0:
                notification_type = NotificationType.STOCK_OUT
                title = f"Sin stock: {saved_product.visible_product_code}"
                body = f"El producto '{saved_product.product_name}' llegó a 0 unidades."
            elif stock_now < saved_product.minimum_stock:
                notification_type = NotificationType.STOCK_LOW
                title = f"Stock bajo: {saved_product.visible_product_code} ({stock_now} unidades)"
                body = f"El producto '{saved_product.product_name}' tiene menos de {saved_product.minimum_stock} unidades disponibles."
            else:
                notification_type = None

            if notification_type is not None:
                deposito_ids = [user.crm_user_id for user in self._user_repository.list_active_by_role_key("deposito")]
                ejecutivo_ids = [user.crm_user_id for user in self._user_repository.list_active_by_role_key("ejecutivo")]
                recipient_ids = list({*deposito_ids, *ejecutivo_ids})
                self._notification_service.notify_bulk(
                    recipient_crm_user_ids=recipient_ids,
                    notification_type=notification_type,
                    title=title,
                    body=body,
                    entity_type=NotificationEntityType.STOCK_PRODUCT,
                    entity_id=saved_product.product_id,
                )

        return saved_product

    def delete_product(self, actor: ResolvedCrmSession, product_id: str) -> StockProduct:
        """Da de baja lógica un producto existente.

        Args:
            actor: Sesión CRM autenticada.
            product_id: Producto a eliminar.

        Returns:
            StockProduct: Producto actualizado tras la baja lógica.
        """

        self._ensure_inventory_admin(actor)
        product = self._get_operable_product(product_id)
        product.deactivate()
        saved = self._product_repository.save(product)
        self._log_event("stock.product_deleted", actor, saved.product_id, saved.product_name, {})
        return saved

    def preview_import(self, actor: ResolvedCrmSession, *, filename: str, content: bytes) -> StockImportPreviewResponse:
        """Parsea y valida una importacion sin tocar el stock real."""

        self._ensure_inventory_write_access(actor)
        rows = self._parse_import_file(filename, content)
        categories = {self._normalize_lookup_key(category.name): category for category in self._category_repository.list_active()}

        seen_codes: set[str] = set()
        parsed_rows = [self._build_import_row(row_number, raw, categories, seen_codes) for row_number, raw in rows]
        if not parsed_rows:
            raise StockImportValidationError("El archivo no contiene filas para importar.")

        batch = StockImportBatch(
            status="pending",
            filename=self._safe_filename(filename),
            created_by_crm_user_id=actor.crm_user.crm_user_id,
        )
        batch.rows.extend(parsed_rows)
        self._refresh_import_batch_counts(batch)

        session = self._product_repository.session
        session.add(batch)
        session.commit()
        session.refresh(batch)

        self._log_event(
            "stock.import_preview",
            actor,
            batch.import_id,
            batch.filename,
            {"total_rows": batch.total_rows, "invalid_rows": batch.invalid_rows},
        )
        return self._build_import_preview_response(batch)

    def confirm_import(self, actor: ResolvedCrmSession, import_id: str) -> StockImportConfirmResponse:
        """Confirma una importacion pendiente dentro de una unica transaccion."""

        self._ensure_inventory_write_access(actor)
        session = self._product_repository.session
        batch = self._get_import_batch(import_id)
        if batch.status != "pending":
            raise StockImportConflictError("La importacion ya fue confirmada, cancelada o restaurada.")
        if batch.invalid_rows:
            raise StockImportValidationError("La importacion tiene filas invalidas y no puede confirmarse.")

        self._revalidate_import_batch(batch)
        warehouse_id = self._product_repository.get_default_warehouse_id()

        try:
            backup = StockBackup(import_id=batch.import_id, created_by_crm_user_id=actor.crm_user.crm_user_id)
            backup.rows.extend(self._build_backup_rows())
            session.add(backup)
            session.flush()

            affected_products: list[StockProduct] = []
            for row in batch.rows:
                product = self._product_repository.get_by_code(row.product_code)
                if product is None:
                    product = StockProduct.create(
                        name=row.product_name,
                        product_code=row.product_code,
                        stock_category_id=row.category_id or "",
                        initial_stock=row.imported_stock,
                        image_url=row.image_url,
                        requires_tracking=False,
                        actor_crm_user_id=actor.crm_user.crm_user_id,
                        warehouse_id=warehouse_id,
                    )
                    product.shelf_id = row.shelf_id
                    product.shelf_height = row.shelf_height
                    session.add(product)
                    session.flush()
                    row.product_id = product.product_id
                else:
                    product.increase_stock(
                        quantity=row.imported_stock,
                        actor_crm_user_id=actor.crm_user.crm_user_id,
                        warehouse_id=warehouse_id,
                        reference_entity_type="stock_import",
                        reference_entity_id=batch.import_id,
                        notes=f"Importacion automatica {batch.filename}",
                    )
                    if row.shelf_id is not None and row.shelf_height is not None:
                        product.shelf_id = row.shelf_id
                        product.shelf_height = row.shelf_height
                    row.product_id = product.product_id
                affected_products.append(product)

            batch.status = "confirmed"
            batch.confirmed_by_crm_user_id = actor.crm_user.crm_user_id
            batch.confirmed_at = datetime.now(UTC)
            session.commit()
        except Exception:
            session.rollback()
            raise

        refreshed_products = [self._product_repository.get_by_id(product.product_id) or product for product in affected_products]
        self._log_event(
            "stock.import_confirmed",
            actor,
            batch.import_id,
            batch.filename,
            {"backup_id": backup.backup_id, "total_import_stock": batch.total_import_stock},
        )
        return StockImportConfirmResponse(
            import_id=batch.import_id,
            backup_id=backup.backup_id,
            status=batch.status,
            created_count=batch.created_count,
            updated_count=batch.updated_count,
            total_import_stock=batch.total_import_stock,
            products=[self._build_product_schema(product) for product in refreshed_products],
        )

    def latest_backup_status(self, actor: ResolvedCrmSession) -> StockBackupStatusResponse:
        """Devuelve el ultimo backup disponible para rollback."""

        self._ensure_inventory_admin(actor)
        backup = self._latest_available_backup()
        if backup is None:
            return StockBackupStatusResponse(has_backup=False)
        return StockBackupStatusResponse(
            has_backup=True,
            import_id=backup.import_id,
            backup_id=backup.backup_id,
            filename=backup.batch.filename if backup.batch else None,
            created_at=backup.created_at,
            total_rows=backup.batch.total_rows if backup.batch else 0,
            total_import_stock=backup.batch.total_import_stock if backup.batch else 0,
        )

    def rollback_import(self, actor: ResolvedCrmSession, import_id: str) -> StockRollbackResponse:
        """Restaura el backup asociado a una importacion confirmada."""

        self._ensure_inventory_admin(actor)
        session = self._product_repository.session
        backup = session.scalar(
            select(StockBackup).where(StockBackup.import_id == import_id, StockBackup.rolled_back_at.is_(None))
        )
        if backup is None:
            raise StockBackupNotFoundError()

        deactivated_created = 0
        restored = 0
        try:
            for backup_row in backup.rows:
                product = self._product_repository.get_by_id(backup_row.product_id)
                if product is None:
                    continue
                product.product_code = backup_row.product_code
                product.product_name = backup_row.product_name
                product.category_id = backup_row.category_id
                product.image_url = backup_row.image_url
                product.shelf_id = backup_row.shelf_id
                product.shelf_height = backup_row.shelf_height
                product.is_active = backup_row.is_active
                if product.primary_stock is None:
                    product.stock_entries.append(
                        self._build_stock_level(quantity=backup_row.current_stock, warehouse_id=self._product_repository.get_default_warehouse_id())
                    )
                else:
                    product.primary_stock.quantity_available = Decimal(backup_row.current_stock)
                restored += 1

            snapshot_codes = {row.product_code for row in backup.rows if row.product_code}
            if backup.batch is not None:
                for import_row in backup.batch.rows:
                    if import_row.product_code in snapshot_codes:
                        continue
                    product = self._product_repository.get_by_code(import_row.product_code)
                    if product is not None and product.is_active:
                        product.deactivate()
                        deactivated_created += 1

            backup.rolled_back_at = datetime.now(UTC)
            backup.rolled_back_by_crm_user_id = actor.crm_user.crm_user_id
            if backup.batch is not None:
                backup.batch.status = "rolled_back"
            session.commit()
        except Exception:
            session.rollback()
            raise

        products = self._product_repository.list_active()
        self._log_event(
            "stock.import_rollback",
            actor,
            backup.import_id,
            backup.batch.filename if backup.batch else backup.backup_id,
            {"backup_id": backup.backup_id, "restored_products": restored},
        )
        return StockRollbackResponse(
            import_id=backup.import_id,
            backup_id=backup.backup_id,
            restored_products=restored,
            deactivated_created_products=deactivated_created,
            products=[self._build_product_schema(product) for product in products],
        )

    def _get_operable_product(self, product_id: str) -> StockProduct:
        """Obtiene un producto existente listo para operar.

        Args:
            product_id: Identificador del producto.

        Returns:
            StockProduct: Producto encontrado.
        """

        product = self._product_repository.get_by_id(product_id)
        if product is None:
            raise StockProductNotFoundError()
        product._ensure_active()
        return product

    def _parse_import_file(self, filename: str, content: bytes) -> list[tuple[int, dict[str, str]]]:
        suffix = PurePath(filename or "").suffix.lower()
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            if reader.fieldnames is None:
                raise StockImportValidationError("El CSV no tiene encabezados.")
            return [(index, self._normalize_import_dict(row)) for index, row in enumerate(reader, start=2)]
        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise StockImportValidationError("El backend no tiene soporte XLSX instalado.") from exc

            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise StockImportValidationError("El XLSX esta vacio.")
            headers = [self._normalize_header(value) for value in rows[0]]
            parsed: list[tuple[int, dict[str, str]]] = []
            for index, values in enumerate(rows[1:], start=2):
                parsed.append((index, self._normalize_import_dict(dict(zip(headers, values, strict=False)))))
            return parsed
        raise StockImportValidationError("El archivo debe ser CSV o XLSX.")

    def _build_import_row(
        self,
        row_number: int,
        raw: dict[str, str],
        categories: dict[str, StockCategory],
        seen_codes: set[str],
    ) -> StockImportRow:
        required = {"imagen", "codigo", "producto", "categoria", "stock", "ubication"}
        errors: list[str] = []
        missing_columns = sorted(required - set(raw))
        if missing_columns:
            errors.append(f"Faltan columnas: {', '.join(missing_columns)}")

        image_url = raw.get("imagen", "").strip()
        product_code = raw.get("codigo", "").strip().upper()
        product_name = raw.get("producto", "").strip()
        category_name = raw.get("categoria", "").strip()
        stock_text = raw.get("stock", "").strip()
        ubication = raw.get("ubication", "").strip()

        for label, value in (
            ("imagen", image_url),
            ("codigo", product_code),
            ("producto", product_name),
            ("categoria", category_name),
            ("stock", stock_text),
        ):
            if not value:
                errors.append(f"{label} es obligatorio.")

        imported_stock = 0
        try:
            stock_decimal = Decimal(stock_text)
            imported_stock = int(stock_decimal)
            if imported_stock <= 0 or stock_decimal != Decimal(imported_stock):
                raise InvalidStockQuantityError()
        except Exception:
            errors.append("stock debe ser un numero entero mayor a cero.")

        category = categories.get(self._normalize_lookup_key(category_name))
        if category is None and category_name:
            errors.append("La categoria no existe o no esta activa.")

        if product_code in seen_codes:
            errors.append("El codigo esta duplicado dentro del archivo.")
        elif product_code:
            seen_codes.add(product_code)

        shelf_id, shelf_height = self._parse_ubication(ubication, errors)
        product = self._product_repository.get_by_code(product_code) if product_code else None
        if product is not None and not product.is_active:
            errors.append("El codigo pertenece a un producto inactivo.")
        old_stock = product.current_stock if product is not None and product.is_active else 0
        is_new_product = product is None

        return StockImportRow(
            row_number=row_number,
            image_url=image_url,
            product_code=product_code,
            product_name=product_name,
            category_name=category_name,
            category_id=category.stock_category_id if category is not None else None,
            imported_stock=imported_stock,
            old_stock=old_stock,
            new_stock=old_stock + imported_stock,
            shelf_id=shelf_id,
            shelf_height=shelf_height,
            is_new_product=is_new_product,
            is_valid=not errors,
            errors=errors,
            product_id=product.product_id if product is not None else None,
        )

    def _revalidate_import_batch(self, batch: StockImportBatch) -> None:
        categories = {category.stock_category_id for category in self._category_repository.list_active()}
        errors: list[str] = []
        for row in batch.rows:
            if row.category_id not in categories:
                errors.append(f"Fila {row.row_number}: la categoria ya no esta activa.")
                continue
            product = self._product_repository.get_by_code(row.product_code)
            if row.is_new_product and product is not None:
                errors.append(f"Fila {row.row_number}: el codigo ahora existe; volve a previsualizar.")
            if not row.is_new_product:
                if product is None or not product.is_active:
                    errors.append(f"Fila {row.row_number}: el producto ya no esta disponible.")
                elif product.current_stock != row.old_stock:
                    errors.append(f"Fila {row.row_number}: el stock cambio desde la previsualizacion.")
        if errors:
            raise StockImportConflictError(" ".join(errors))

    def _build_backup_rows(self) -> list[StockBackupRow]:
        rows: list[StockBackupRow] = []
        for product in self._product_repository.list_all():
            rows.append(
                StockBackupRow(
                    product_id=product.product_id,
                    product_code=product.product_code,
                    product_name=product.product_name,
                    category_id=product.category_id,
                    image_url=product.image_url,
                    current_stock=product.current_stock,
                    shelf_id=product.shelf_id,
                    shelf_height=product.shelf_height,
                    is_active=product.is_active,
                )
            )
        return rows

    def _get_import_batch(self, import_id: str) -> StockImportBatch:
        batch = self._product_repository.session.get(StockImportBatch, import_id)
        if batch is None:
            raise StockImportNotFoundError()
        return batch

    def _latest_available_backup(self) -> StockBackup | None:
        return self._product_repository.session.scalar(
            select(StockBackup)
            .where(StockBackup.rolled_back_at.is_(None))
            .order_by(StockBackup.created_at.desc())
            .limit(1)
        )

    def _refresh_import_batch_counts(self, batch: StockImportBatch) -> None:
        rows = list(batch.rows)
        valid_rows = [row for row in rows if row.is_valid]
        batch.total_rows = len(rows)
        batch.valid_rows = len(valid_rows)
        batch.invalid_rows = len(rows) - len(valid_rows)
        batch.created_count = sum(1 for row in valid_rows if row.is_new_product)
        batch.updated_count = sum(1 for row in valid_rows if not row.is_new_product)
        batch.total_import_stock = sum(row.imported_stock for row in valid_rows)

    def _build_import_preview_response(self, batch: StockImportBatch) -> StockImportPreviewResponse:
        return StockImportPreviewResponse(
            import_id=batch.import_id,
            status=batch.status,
            filename=batch.filename,
            total_rows=batch.total_rows,
            valid_rows=batch.valid_rows,
            invalid_rows=batch.invalid_rows,
            created_count=batch.created_count,
            updated_count=batch.updated_count,
            total_import_stock=batch.total_import_stock,
            can_confirm=batch.status == "pending" and batch.invalid_rows == 0 and batch.valid_rows > 0,
            rows=[
                StockImportRowPreview(
                    row_number=row.row_number,
                    image_url=row.image_url,
                    product_code=row.product_code,
                    product_name=row.product_name,
                    category_name=row.category_name,
                    imported_stock=row.imported_stock,
                    old_stock=row.old_stock,
                    new_stock=row.new_stock,
                    ubication=self._format_ubication(row.shelf_id, row.shelf_height),
                    shelf_id=row.shelf_id,
                    shelf_height=row.shelf_height,
                    is_new_product=row.is_new_product,
                    is_valid=row.is_valid,
                    errors=list(row.errors or []),
                )
                for row in batch.rows
            ],
        )

    def _build_product_schema(self, product: StockProduct) -> StockProductResponse:
        return StockProductResponse(
            id=product.stock_product_id,
            product_id=product.stock_product_id,
            product_code=product.visible_product_code,
            name=product.name,
            product_name=product.name,
            category_id=product.stock_category_id or "",
            category_name=product.category.name,
            current_stock=product.current_stock,
            image_url=product.image_url,
            minimum_stock=product.minimum_stock,
            shelf_id=product.shelf_id,
            shelf_height=product.shelf_height,
            requires_tracking=product.requires_tracking,
            created_at=product.created_at,
            updated_at=product.updated_at,
            is_active=product.is_active,
        )

    def _build_stock_level(self, *, quantity: int, warehouse_id: str) -> StockLevel:
        return StockLevel.create(warehouse_id=warehouse_id, quantity_available=quantity)

    def _parse_ubication(self, value: str, errors: list[str]) -> tuple[str | None, int | None]:
        if not value:
            return None, None
        match = re.fullmatch(r"\s*([A-Za-z])\s*[-/]\s*(\d+)\s*", value)
        if match is None:
            errors.append("ubication debe tener formato A-3 o A/3.")
            return None, None
        height = int(match.group(2))
        if height < 1:
            errors.append("ubication debe indicar una altura mayor a cero.")
            return None, None
        return match.group(1).upper(), height

    def _normalize_import_dict(self, row: dict[object, object]) -> dict[str, str]:
        return {self._normalize_header(key): "" if value is None else str(value).strip() for key, value in row.items() if key is not None}

    def _normalize_header(self, value: object) -> str:
        return str(value or "").strip().lower()

    def _normalize_lookup_key(self, value: str) -> str:
        return re.sub(r"\s+", " ", value.strip()).lower()

    def _format_ubication(self, shelf_id: str | None, shelf_height: int | None) -> str | None:
        if shelf_id is None or shelf_height is None:
            return None
        return f"{shelf_id}-{shelf_height}"

    def _safe_filename(self, filename: str) -> str:
        return PurePath(filename or "stock-import").name[:255]

    def _ensure_inventory_read_access(self, actor: ResolvedCrmSession) -> None:
        """Valida que la sesión tenga acceso de lectura al módulo real de depósito.

        Args:
            actor: Sesión CRM autenticada.
        """

        if "admin" in actor.role_keys:
            return

        membership = actor.auth_result.active_membership
        if not {"deposito", "ejecutivo"}.intersection(actor.role_keys):
            raise InventoryAccessDeniedError()
        if membership.tenant_type != "company" or membership.tenant_id != self._settings.deposito_demo_tenant_id:
            raise InventoryAccessDeniedError()

    def _ensure_inventory_write_access(self, actor: ResolvedCrmSession) -> None:
        """Valida que la sesión tenga permisos operativos sobre depósito.

        Args:
            actor: Sesión CRM autenticada.
        """

        self._ensure_inventory_read_access(actor)
        if self._permission_service is None:
            if "deposito" in actor.role_keys or "admin" in actor.role_keys:
                return
            raise InventoryAccessDeniedError()
        if not self._permission_service.resolve(actor.role_keys, actor.crm_user.crm_user_id, PERMISSION_STOCK_MANAGE):
            raise InventoryAccessDeniedError()

    def _ensure_inventory_admin(self, actor: ResolvedCrmSession) -> None:
        """Valida que la sesión tenga permisos administrativos en el CRM."""

        if self._permission_service is None:
            if "admin" in actor.role_keys:
                return
            raise InventoryAdminRequiredError()
        if not self._permission_service.resolve(actor.role_keys, actor.crm_user.crm_user_id, PERMISSION_STOCK_DELETE_PRODUCT):
            raise InventoryAdminRequiredError()

    def _log_event(
        self,
        event_code: str,
        actor: ResolvedCrmSession,
        entity_id: str,
        entity_label: str,
        extra: dict[str, object],
    ) -> None:
        if self._activity_log_service is None:
            return
        self._activity_log_service.log(
            event_code,
            actor,
            entity_type="stock_product",
            entity_id=entity_id,
            entity_label=entity_label,
            summary=event_code,
            extra=extra,
        )
